# utils/report_generator.py
import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

class ReportGenerator:
    """
    Gerador de relatórios em Excel com formatação profissional.
    """
    
    def __init__(self, analysis_results):
        """
        Inicializa o gerador de relatórios.
        
        Args:
            analysis_results: Dicionário com resultados da análise
        """
        self.results = analysis_results
        self.wb = None
        
        # Estilos Profarma
        self.header_fill = PatternFill(start_color="14555a", end_color="14555a", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.subheader_fill = PatternFill(start_color="00aeef", end_color="00aeef", fill_type="solid")
        self.subheader_font = Font(color="FFFFFF", bold=True, size=11)
        self.outlier_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_excel_report(self, output_dir='reports'):
        """
        Gera relatório completo em Excel.
        
        Args:
            output_dir: Diretório de saída
            
        Returns:
            str: Caminho do arquivo gerado
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # Nome do arquivo com timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'relatorio_sazonalidade_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            # Criar workbook
            self.wb = Workbook()
            
            # Remover aba padrão
            default_sheet = self.wb.active
            self.wb.remove(default_sheet)
            
            # Criar abas
            self._create_summary_sheet()
            self._create_details_sheet()
            self._create_outliers_sheet()
            self._create_statistics_sheet()
            
            # Salvar arquivo
            self.wb.save(filepath)
            
            print(f"Relatório gerado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Erro ao gerar relatório: {str(e)}")
            return None
    
    def _create_summary_sheet(self):
        """Cria aba de resumo executivo."""
        ws = self.wb.create_sheet("Resumo Executivo")
        
        # Título
        ws['A1'] = "RELATÓRIO DE ANÁLISE DE SAZONALIDADE"
        ws['A1'].font = Font(size=16, bold=True, color="14555a")
        ws.merge_cells('A1:F1')
        
        # Data do relatório
        ws['A3'] = "Data do Relatório:"
        ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # Parâmetros utilizados
        ws['A5'] = "PARÂMETROS DA ANÁLISE"
        ws['A5'].font = self.header_font
        ws['A5'].fill = self.header_fill
        ws.merge_cells('A5:B5')
        
        ws['A6'] = "Sigma Threshold:"
        ws['B6'] = self.results.get('summary', {}).get('sigma_threshold', 2)
        
        # Estatísticas gerais
        ws['A8'] = "ESTATÍSTICAS GERAIS"
        ws['A8'].font = self.header_font
        ws['A8'].fill = self.header_fill
        ws.merge_cells('A8:B8')
        
        summary = self.results.get('summary', {})
        ws['A9'] = "Total de Categorias:"
        ws['B9'] = summary.get('processed', 0)
        ws['A10'] = "Processadas com Sucesso:"
        ws['B10'] = summary.get('successful', 0)
        ws['A11'] = "Categorias com Erro:"
        ws['B11'] = summary.get('errors', 0)
        
        # Formatar colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Adicionar bordas
        for row in range(5, 12):
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.border = self.border
    
    def _create_details_sheet(self):
        """Cria aba com detalhes de todas as categorias."""
        ws = self.wb.create_sheet("Detalhes por Categoria")
        
        # Cabeçalhos
        headers = [
            'Categoria',
            'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
            'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez',
            'Total Outliers'
        ]
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Dados
        row_num = 2
        categories = self.results.get('categories', [])
        
        for cat in categories:
            # Nome da categoria
            ws.cell(row=row_num, column=1, value=cat.get('category', ''))
            
            # Sazonalidade por mês
            seasonality = cat.get('seasonality_year', [])
            for month_idx, value in enumerate(seasonality[:12], 2):
                cell = ws.cell(row=row_num, column=month_idx, value=value)
                cell.number_format = '0.00%'
                
                # Destacar outliers
                outliers = cat.get('outliers', {}).get('months', [])
                if month_idx - 2 in outliers:
                    cell.fill = self.outlier_fill
            
            # Total de outliers
            total_outliers = len(cat.get('outliers', {}).get('months', []))
            ws.cell(row=row_num, column=14, value=total_outliers)
            
            row_num += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        for col in range(2, 15):
            ws.column_dimensions[chr(64 + col)].width = 10
        
        # Adicionar bordas
        for row in range(1, row_num):
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = self.border
    
    def _create_outliers_sheet(self):
        """Cria aba específica para outliers."""
        ws = self.wb.create_sheet("Análise de Outliers")
        
        # Título
        ws['A1'] = "ANÁLISE DE OUTLIERS POR CATEGORIA"
        ws['A1'].font = Font(size=14, bold=True, color="14555a")
        ws.merge_cells('A1:E1')
        
        # Cabeçalhos
        headers = ['Categoria', 'Mês', 'Valor Original', 'Desvio (σ)', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Dados de outliers
        row_num = 4
        month_names = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                       'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        for cat in self.results.get('categories', []):
            outliers = cat.get('outliers', {})
            outlier_months = outliers.get('months', [])
            outlier_values = outliers.get('values', [])
            
            for month, value in zip(outlier_months, outlier_values):
                ws.cell(row=row_num, column=1, value=cat.get('category', ''))
                ws.cell(row=row_num, column=2, value=month_names[month] if month < 12 else str(month))
                ws.cell(row=row_num, column=3, value=value).number_format = '0.00'
                ws.cell(row=row_num, column=4, value=abs(value))
                ws.cell(row=row_num, column=5, value='Outlier Detectado')
                
                # Colorir linha
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = self.outlier_fill
                    ws.cell(row=row_num, column=col).border = self.border
                
                row_num += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
    
    def _create_statistics_sheet(self):
        """Cria aba com estatísticas detalhadas."""
        ws = self.wb.create_sheet("Estatísticas")
        
        # Título
        ws['A1'] = "ESTATÍSTICAS DETALHADAS POR CATEGORIA"
        ws['A1'].font = Font(size=14, bold=True, color="14555a")
        ws.merge_cells('A1:G1')
        
        # Cabeçalhos
        headers = [
            'Categoria', 
            'Média', 
            'Desvio Padrão',
            'CV%',
            'Min',
            'Max',
            'Amplitude'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Dados
        row_num = 4
        for cat in self.results.get('categories', []):
            stats = cat.get('statistics', {})
            
            ws.cell(row=row_num, column=1, value=cat.get('category', ''))
            ws.cell(row=row_num, column=2, value=stats.get('mean', 0)).number_format = '0.00'
            ws.cell(row=row_num, column=3, value=stats.get('std', 0)).number_format = '0.00'
            ws.cell(row=row_num, column=4, value=stats.get('cv', 0)).number_format = '0.00%'
            ws.cell(row=row_num, column=5, value=stats.get('min', 0)).number_format = '0.00'
            ws.cell(row=row_num, column=6, value=stats.get('max', 0)).number_format = '0.00'
            ws.cell(row=row_num, column=7, value=stats.get('range', 0)).number_format = '0.00'
            
            # Bordas
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = self.border
            
            row_num += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def generate_json_report(self, output_dir='reports'):
        """
        Gera relatório em formato JSON.
        
        Args:
            output_dir: Diretório de saída
            
        Returns:
            str: Caminho do arquivo gerado
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'relatorio_sazonalidade_{timestamp}.json'
            filepath = os.path.join(output_dir, filename)
            
            # Converter arrays numpy para listas
            def convert_to_serializable(obj):
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                elif isinstance(obj, np.integer):
                    return int(obj)
                elif isinstance(obj, np.floating):
                    if np.isnan(obj) or np.isinf(obj):
                        return None
                    return float(obj)
                elif isinstance(obj, dict):
                    return {k: convert_to_serializable(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [convert_to_serializable(v) for v in obj]
                return obj
            
            serializable_results = convert_to_serializable(self.results)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(serializable_results, f, indent=2, ensure_ascii=False)
            
            print(f"Relatório JSON gerado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Erro ao gerar relatório JSON: {str(e)}")
            return None
